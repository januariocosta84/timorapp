import os
import django

# Set the DJANGO_SETTINGS_MODULE environment variable to your settings file
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'TimorWebApp.settings')

# Initialize Django
django.setup()
from openpyxl import load_workbook
from parentregistrationApp.models import Municipality, AdministrativePost, Suco, Village


file_path = 'munic.xlsx'


workbook = load_workbook(file_path)
sheet = workbook.active  

for row in sheet.iter_rows(min_row=2, values_only=True):
    municipality_name = row[1]  # Column 1 (Municipality)
    administrative_post_name = row[2]  # Column 2 (Administrative Post)
    suco_name = row[3]  # Column 3 (Suco)
    village_name = row[4]  # Column 4 (Village)

    print(municipality_name)
    # Check if the Municipality exists, otherwise create it
    municipality, created = Municipality.objects.get_or_create(name=municipality_name)

    # Check if the Administrative Post exists, otherwise create it
    administrative_post, created = AdministrativePost.objects.get_or_create(
        name=administrative_post_name,
        municipality=municipality  # Link it to the Municipality
    )

    # Check if the Suco exists, otherwise create it
    suco, created = Suco.objects.get_or_create(
        name=suco_name,
        posto=administrative_post  # Link it to the Administrative Post
    )

    # Check if the Village exists, otherwise create it
    village, created = Village.objects.get_or_create(
        name=village_name,
        suco=suco  # Link it to the Suco
    )

    # Optionally print the created/updated records for verification
    print(f"Created: Municipality: {municipality.name}, Administrative Post: {administrative_post.name}, Suco: {suco.name}, Village: {village.name}")
